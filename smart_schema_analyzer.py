"""
Smart Schema Analyzer
Analyzes Excel schema files and detects data types, constraints, and NDMO compliance
"""

import pandas as pd
import os
from datetime import datetime

class SmartSchemaAnalyzer:
    """Smart Schema Analyzer for NDMO Compliance"""
    
    def __init__(self):
        self.schema_data = None
        self.analysis_results = {}
    
    def analyze_schema(self, schema_file_path):
        """Analyze schema file and return comprehensive analysis"""
        try:
            # Read schema file
            if not os.path.exists(schema_file_path):
                return {'error': f'Schema file not found: {schema_file_path}'}
            
            # Check file size
            file_size = os.path.getsize(schema_file_path)
            if file_size == 0:
                return {'error': 'Schema file is empty'}
            
            # CRITICAL: Use openpyxl directly to get ALL columns (including empty ones)
            import openpyxl
            df = None
            actual_column_count = 0
            column_names = []
            
            try:
                # Step 1: Read Excel file with openpyxl to get exact column count
                # Try to load workbook with error handling
                try:
                    wb = openpyxl.load_workbook(schema_file_path, read_only=True, data_only=True)
                except Exception as wb_error:
                    # Try without read_only if that fails
                    try:
                        wb = openpyxl.load_workbook(schema_file_path, data_only=True)
                    except Exception as wb_error2:
                        return {'error': f'Cannot read Excel file. File may be corrupted or in wrong format. Error: {str(wb_error)} / {str(wb_error2)}'}
                ws = wb.active
                
                # Get actual maximum column from Excel file
                max_col = ws.max_column
                actual_column_count = max_col
                
                # Get column names from first row
                for col_idx in range(1, max_col + 1):
                    cell_value = ws.cell(row=1, column=col_idx).value
                    if cell_value and str(cell_value).strip():
                        column_names.append(str(cell_value).strip())
                    else:
                        column_names.append(f'Column_{col_idx}')
                
                wb.close()
                
                # Step 2: Read with pandas using explicit column range to get ALL columns
                try:
                    # CRITICAL: Read with usecols to force reading ALL columns
                    df = pd.read_excel(
                        schema_file_path, 
                        engine='openpyxl', 
                        header=0,
                        usecols=range(max_col)  # Force read all columns
                    )
                    
                    # Ensure we have exactly max_col columns
                    if len(df.columns) < max_col:
                        # Add missing columns
                        for i in range(len(df.columns), max_col):
                            col_name = column_names[i] if i < len(column_names) else f'Column_{i+1}'
                            df[col_name] = None
                    
                    # Rename columns to match what we found
                    if len(column_names) == len(df.columns):
                        df.columns = column_names
                    elif len(column_names) > 0:
                        # Use column names we found, pad if needed
                        new_columns = []
                        for i in range(max_col):
                            if i < len(column_names):
                                new_columns.append(column_names[i])
                            elif i < len(df.columns):
                                new_columns.append(df.columns[i])
                            else:
                                new_columns.append(f'Column_{i+1}')
                        df.columns = new_columns[:len(df.columns)]
                        
                except Exception as e_read:
                    # Fallback: read without header and manually set columns
                    try:
                        df = pd.read_excel(schema_file_path, engine='openpyxl', header=None)
                        # Check if first row looks like headers
                        if len(df) > 0:
                            first_row = df.iloc[0]
                            non_null_count = first_row.notna().sum()
                            
                            if non_null_count > max_col * 0.3:
                                # First row is headers
                                new_columns = []
                                for i in range(max_col):
                                    if i < len(first_row):
                                        val = first_row.iloc[i]
                                        if pd.notna(val):
                                            new_columns.append(str(val).strip())
                                        else:
                                            new_columns.append(column_names[i] if i < len(column_names) else f'Column_{i+1}')
                                    else:
                                        new_columns.append(column_names[i] if i < len(column_names) else f'Column_{i+1}')
                                
                                df.columns = new_columns[:len(df.columns)]
                                df = df.iloc[1:].reset_index(drop=True)
                            else:
                                # No clear headers, use column names we found
                                new_columns = column_names[:max_col] if len(column_names) >= max_col else column_names + [f'Column_{i+1}' for i in range(len(column_names), max_col)]
                                df.columns = new_columns[:len(df.columns)]
                            
                            # Ensure we have max_col columns
                            if len(df.columns) < max_col:
                                for i in range(len(df.columns), max_col):
                                    col_name = column_names[i] if i < len(column_names) else f'Column_{i+1}'
                                    df[col_name] = None
                        else:
                            # Empty file, create columns
                            df = pd.DataFrame(columns=column_names[:max_col] if len(column_names) >= max_col else column_names + [f'Column_{i+1}' for i in range(len(column_names), max_col)])
                    except Exception as e2:
                        raise Exception(f"Failed to read Excel: {str(e_read)} / {str(e2)}")
                
            except Exception as e1:
                try:
                    # Fallback: try reading with pandas default
                    df = pd.read_excel(schema_file_path, engine='openpyxl')
                    actual_column_count = len(df.columns)
                    # Try to get more columns
                    try:
                        df_test = pd.read_excel(schema_file_path, header=None, engine='openpyxl', nrows=1)
                        actual_column_count = max(actual_column_count, len(df_test.columns))
                    except:
                        pass
                except Exception as e2:
                    return {'error': f'Failed to read schema file: {str(e1)} / {str(e2)}'}
            
            if df is None or df.empty:
                return {'error': 'Schema file is empty or could not be read.'}
            
            # Ensure we have all columns
            if len(df.columns) < actual_column_count:
                # Add missing columns
                for i in range(len(df.columns), actual_column_count):
                    col_name = column_names[i] if i < len(column_names) else f'Column_{i+1}'
                    df[col_name] = None
            
            # Analyze schema structure
            analysis = {
                'timestamp': datetime.now().isoformat(),
                'file_name': os.path.basename(schema_file_path),
                'total_fields': len(df),
                'total_columns': max(len(df.columns), actual_column_count),
                'columns': list(df.columns),
                'fields': [],
                'has_primary_key': False,
                'has_foreign_keys': False,
                'has_audit_trail': False,
                'data_types': {},
                'constraints': [],
                'issues': [],
                'recommendations': [],
                'ndmo_compliance': {}
            }
            
            # Analyze each column/field
            for idx, row in df.iterrows():
                field_info = self._analyze_field(row, df.columns)
                analysis['fields'].append(field_info)
                
                # Check for primary key
                if field_info.get('is_primary_key', False):
                    analysis['has_primary_key'] = True
                
                # Check for foreign keys
                if field_info.get('is_foreign_key', False):
                    analysis['has_foreign_keys'] = True
                
                # Check for audit trail fields
                if field_info.get('is_audit_field', False):
                    analysis['has_audit_trail'] = True
                
                # Track data types
                data_type = field_info.get('detected_type', 'Unknown')
                analysis['data_types'][data_type] = analysis['data_types'].get(data_type, 0) + 1
            
            # Analyze each column for NDMO compliance - THIS IS THE KEY PART
            analysis['column_analysis'] = self._analyze_columns(df)
            
            # CRITICAL: Ensure we analyze ALL columns from Excel file
            # If we have fewer analyzed columns than actual, analyze remaining columns
            if len(analysis['column_analysis']) < actual_column_count:
                # Analyze remaining columns that might have been missed
                for col_idx in range(len(df.columns), actual_column_count):
                    col_info = {
                        'column_name': f'Column_{col_idx + 1}',
                        'data_type': 'Unknown',
                        'non_null_count': 0,
                        'null_count': len(df) if len(df) > 0 else 0,
                        'unique_count': 0,
                        'total_count': len(df),
                        'completeness': 0.0,
                        'uniqueness': 0.0,
                        'detected_type': 'Unknown',
                        'is_primary_key': False,
                        'is_audit_field': False,
                        'ndmo_standards': []
                    }
                    analysis['column_analysis'].append(col_info)
            
            # Update total_columns based on actual analysis
            analysis['total_columns'] = max(len(analysis['column_analysis']), actual_column_count)
            
            # Calculate NDMO compliance per column
            analysis['ndmo_compliance'] = self._calculate_ndmo_compliance_per_column(analysis)
            
            # Generate recommendations
            analysis['recommendations'] = self._generate_recommendations(analysis)
            
            # Identify issues
            analysis['issues'] = self._identify_issues(analysis)
            
            self.schema_data = df
            self.analysis_results = analysis
            
            return analysis
            
        except Exception as e:
            import traceback
            return {'error': f'Error analyzing schema: {str(e)}\n{traceback.format_exc()}'}
    
    def _analyze_field(self, row, columns):
        """Analyze individual field"""
        # Try to get field name from common column names
        field_name = None
        for col_name in ['Field Name', 'Column Name', 'Field', 'Column', 'Name']:
            if col_name in columns:
                field_name = str(row.get(col_name, ''))
                break
        
        if not field_name:
            # Use first column or index
            field_name = str(row.get(columns[0], f'Field_{row.name}'))
        
        field_info = {
            'field_name': field_name,
            'detected_type': 'Text',
            'is_primary_key': False,
            'is_foreign_key': False,
            'is_audit_field': False,
            'is_required': False,
            'constraints': [],
            'business_rules': [],
            'ndmo_standards': [],
            'compliance_score': 0.0
        }
        
        # Detect field name
        field_name_lower = field_name.lower()
        
        # Check for primary key
        pk_keywords = ['id', 'key', 'pk', 'primary', '_id', 'identifier']
        if any(keyword in field_name_lower for keyword in pk_keywords):
            field_info['is_primary_key'] = True
            field_info['ndmo_standards'].append('DG001')  # Unique Identifiers
        
        # Check for foreign keys
        fk_keywords = ['foreign', 'fk', 'reference', 'ref', '_ref', '_fk']
        if any(keyword in field_name_lower for keyword in fk_keywords):
            field_info['is_foreign_key'] = True
            field_info['ndmo_standards'].append('BR002')  # Data Relationships
        
        # Check for audit trail fields
        audit_keywords = ['created', 'updated', 'modified', 'deleted', 'timestamp', 'date', 'user', 'audit', 'by', 'at', 'on']
        if any(keyword in field_name_lower for keyword in audit_keywords):
            field_info['is_audit_field'] = True
            field_info['ndmo_standards'].append('DS004')  # Audit Trail
        
        # Detect data type from field name or row data
        detected_type = self._detect_data_type_from_name(field_name_lower)
        
        # Try to get type from columns
        for type_col_name in ['Type', 'Data Type', 'DataType', 'Field Type']:
            if type_col_name in columns:
                type_value = str(row.get(type_col_name, '')).lower()
                if type_value:
                    detected_type = self._detect_data_type(type_value)
                    break
        
        field_info['detected_type'] = detected_type
        
        # Add data quality standards based on type
        if detected_type == 'Email':
            field_info['ndmo_standards'].append('DQ005')  # Data Validity
        elif detected_type == 'Phone':
            field_info['ndmo_standards'].append('DQ005')  # Data Validity
        elif detected_type == 'DateTime':
            field_info['ndmo_standards'].append('DQ006')  # Data Timeliness
        
        # Check if required
        for req_col_name in ['Required', 'Mandatory', 'Nullable']:
            if req_col_name in columns:
                req_value = str(row.get(req_col_name, '')).lower()
                if req_value in ['yes', 'true', '1', 'y', 'required', 'mandatory']:
                    field_info['is_required'] = True
                    field_info['ndmo_standards'].append('DQ001')  # Data Completeness
                elif req_value in ['no', 'false', '0', 'n', 'optional', 'nullable']:
                    field_info['is_required'] = False
                break
        
        # Calculate compliance score for this field
        field_info['compliance_score'] = self._calculate_field_compliance_score(field_info)
        
        return field_info
    
    def _detect_data_type_from_name(self, field_name):
        """Detect data type from field name"""
        field_name = field_name.lower()
        
        if any(t in field_name for t in ['email', 'mail', 'e-mail']):
            return 'Email'
        elif any(t in field_name for t in ['phone', 'mobile', 'tel', 'telephone']):
            return 'Phone'
        elif any(t in field_name for t in ['date', 'time', 'datetime', 'timestamp', 'created', 'updated', 'modified']):
            return 'DateTime'
        elif any(t in field_name for t in ['id', 'number', 'num', 'count', 'quantity', 'amount', 'price', 'cost']):
            return 'Numeric'
        elif any(t in field_name for t in ['flag', 'is_', 'has_', 'active', 'enabled', 'status']):
            return 'Boolean'
        else:
            return 'Text'
    
    def _calculate_field_compliance_score(self, field_info):
        """Calculate NDMO compliance score for a field"""
        score = 0.0
        max_score = 0.0
        
        # Primary key check (DG001)
        max_score += 0.15
        if field_info.get('is_primary_key'):
            score += 0.15
        
        # Required field check (DQ001)
        max_score += 0.15
        if field_info.get('is_required'):
            score += 0.15
        
        # Audit trail check (DS004)
        max_score += 0.10
        if field_info.get('is_audit_field'):
            score += 0.10
        
        # Data type detection (DQ005)
        max_score += 0.10
        if field_info.get('detected_type') != 'Unknown':
            score += 0.10
        
        return score / max_score if max_score > 0 else 0.0
    
    def _analyze_columns(self, df):
        """Analyze all columns in the dataframe - CRITICAL: Analyze ALL columns"""
        column_analysis = []
        
        # Get all columns from the dataframe
        all_columns = list(df.columns)
        
        # Analyze each column - iterate through ALL columns
        for col_idx, col in enumerate(all_columns):
            try:
                # Get column name
                if isinstance(col, (int, float)):
                    col_name = f'Column_{int(col)+1}'
                else:
                    col_name = str(col).strip() if col else f'Column_{col_idx+1}'
                
                # Get column data - ensure we get the data
                try:
                    if col in df.columns:
                        col_data = df[col]
                    elif col_idx < len(df.columns):
                        col_data = df.iloc[:, col_idx]
                    else:
                        # Column doesn't exist - create empty series
                        col_data = pd.Series([None] * len(df), name=col_name, dtype=object)
                except Exception as e_col:
                    # Create empty series if column doesn't exist
                    col_data = pd.Series([None] * len(df), name=col_name, dtype=object)
                
                # Analyze even if column is empty - we still want to record it
                # Calculate basic stats
                total_count = len(col_data) if len(col_data) > 0 else len(df)
                non_null_count = int(col_data.notna().sum()) if len(col_data) > 0 else 0
                null_count = int(col_data.isna().sum()) if len(col_data) > 0 else total_count
                unique_count = int(col_data.nunique()) if non_null_count > 0 else 0
                
                # Calculate completeness
                completeness = (non_null_count / total_count * 100) if total_count > 0 else 0.0
                
                # Calculate uniqueness
                uniqueness = (unique_count / non_null_count * 100) if non_null_count > 0 else 0.0
                
                # Detect data type
                detected_type = 'Unknown'
                if non_null_count > 0:
                    try:
                        sample_values = col_data.dropna().head(10)
                        detected_type = self._detect_data_type_from_name(col_name.lower())
                        # Try to detect from actual values
                        if len(sample_values) > 0:
                            first_val = sample_values.iloc[0] if hasattr(sample_values, 'iloc') else sample_values[0]
                            if pd.api.types.is_numeric_dtype(col_data):
                                detected_type = 'Numeric'
                            elif pd.api.types.is_datetime64_any_dtype(col_data):
                                detected_type = 'DateTime'
                            elif isinstance(first_val, str) and '@' in str(first_val):
                                detected_type = 'Email'
                            elif isinstance(first_val, str) and any(c.isdigit() for c in str(first_val)) and len(str(first_val)) >= 10:
                                detected_type = 'Phone'
                    except:
                        detected_type = self._detect_data_type_from_name(col_name.lower())
                else:
                    # Empty column - detect from name
                    detected_type = self._detect_data_type_from_name(col_name.lower())
                
                # Build column info
                col_info = {
                    'column_name': col_name,
                    'data_type': str(col_data.dtype) if len(col_data) > 0 else 'Unknown',
                    'non_null_count': non_null_count,
                    'null_count': null_count,
                    'unique_count': unique_count,
                    'total_count': total_count,
                    'completeness': completeness,
                    'uniqueness': uniqueness,
                    'detected_type': detected_type,
                    'is_primary_key': False,
                    'is_audit_field': False,
                    'ndmo_standards': []
                }
                
                # Check for primary key
                col_info['is_primary_key'] = any(kw in col_name.lower() for kw in ['id', 'key', 'pk', 'primary', '_id', 'serial'])
                
                # Check for audit fields
                col_info['is_audit_field'] = any(kw in col_name.lower() for kw in ['created', 'updated', 'modified', 'deleted', 'timestamp', 'date', 'user', 'audit', 'created_by', 'updated_by', 'modified_by'])
                
                # NDMO standards applicable
                if col_info['is_primary_key']:
                    col_info['ndmo_standards'].append('DG001')
                if col_info['is_audit_field']:
                    col_info['ndmo_standards'].append('DS004')
                if col_info['completeness'] < 100:
                    col_info['ndmo_standards'].append('DQ001')
                if col_info['uniqueness'] < 100 and not col_info['is_primary_key']:
                    col_info['ndmo_standards'].append('DQ004')
                
                column_analysis.append(col_info)
            except Exception as e:
                # Log error but continue with other columns
                print(f"Error analyzing column {col}: {str(e)}")
                continue
        
        return column_analysis
    
    def _calculate_ndmo_compliance_per_column(self, analysis):
        """Calculate NDMO compliance for each column"""
        compliance = {}
        
        for col_info in analysis.get('column_analysis', []):
            col_name = col_info['column_name']
            score = 0.0
            max_score = 0.0
            
            # DG001: Unique Identifiers
            if col_info.get('is_primary_key'):
                max_score += 0.15
                score += 0.15
            
            # DQ001: Data Completeness
            max_score += 0.15
            completeness = col_info.get('completeness', 0)
            score += 0.15 * (completeness / 100)
            
            # DQ004: Data Uniqueness
            max_score += 0.10
            uniqueness = col_info.get('uniqueness', 0)
            score += 0.10 * (uniqueness / 100)
            
            # DS004: Audit Trail
            if col_info.get('is_audit_field'):
                max_score += 0.10
                score += 0.10
            
            # DQ005: Data Validity (type detection)
            max_score += 0.10
            if col_info.get('detected_type') != 'Text':
                score += 0.10
            
            compliance[col_name] = {
                'score': score / max_score if max_score > 0 else 0.0,
                'standards': col_info.get('ndmo_standards', []),
                'completeness': completeness,
                'uniqueness': uniqueness,
                'data_type': col_info.get('detected_type', 'Unknown')
            }
        
        return compliance
    
    def _detect_data_type(self, type_str):
        """Detect data type from string"""
        type_str = type_str.lower()
        
        if any(t in type_str for t in ['int', 'integer', 'number', 'numeric']):
            return 'Numeric'
        elif any(t in type_str for t in ['date', 'time', 'datetime', 'timestamp']):
            return 'DateTime'
        elif any(t in type_str for t in ['email', 'mail']):
            return 'Email'
        elif any(t in type_str for t in ['phone', 'mobile', 'tel']):
            return 'Phone'
        elif any(t in type_str for t in ['bool', 'boolean', 'yes/no']):
            return 'Boolean'
        else:
            return 'Text'
    
    def _generate_recommendations(self, analysis):
        """Generate recommendations based on analysis"""
        recommendations = []
        
        if not analysis['has_primary_key']:
            recommendations.append({
                'type': 'Critical',
                'message': 'Add primary key field to ensure data uniqueness (DG001)',
                'standard': 'DG001'
            })
        
        if not analysis['has_audit_trail']:
            recommendations.append({
                'type': 'Critical',
                'message': 'Add audit trail fields (created_date, updated_date, created_by) (DS004)',
                'standard': 'DS004'
            })
        
        if analysis['has_foreign_keys']:
            recommendations.append({
                'type': 'Info',
                'message': 'Foreign key relationships detected. Ensure referential integrity is maintained.',
                'standard': 'BR002'
            })
        
        return recommendations
    
    def _identify_issues(self, analysis):
        """Identify schema issues"""
        issues = []
        
        if not analysis['has_primary_key']:
            issues.append({
                'severity': 'High',
                'issue': 'Missing Primary Key',
                'impact': 'Cannot ensure data uniqueness',
                'standard': 'DG001'
            })
        
        if not analysis['has_audit_trail']:
            issues.append({
                'severity': 'High',
                'issue': 'Missing Audit Trail',
                'impact': 'Cannot track data changes',
                'standard': 'DS004'
            })
        
        return issues
    
    def get_analysis_results(self):
        """Get stored analysis results"""
        return self.analysis_results

